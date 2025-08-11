import pandas as pd
import openpyxl
from openpyxl.cell.cell import MergedCell
from io import BytesIO
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment

def add_full_border(ws, min_row, max_row, min_col, max_col):
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            
def fill_template_excel(template_file, dcl_df, csv_df, dcl_data, dat_data=None):
    wb = openpyxl.load_workbook(template_file)
    ws_result = wb['Test Result']
    ws_parts = wb['PartsCoverage']

    # 写入Test Result（dcl数据）
    ws_result['I7'] = dcl_data.get("passfail", "")
    ws_result['I8'] = dcl_data.get("test_time", "")
    ws_result['I4'] = dcl_data.get("board_name", "")

    # 获取Test Result表头（假设表头在第10行，数据从A11开始）
    header_row = 10
    start_row = 11
    headers = [cell.value for cell in ws_result[header_row]]

    # 保证dcl_df列顺序与表头一致，只保留表头列
    dcl_df_to_write = dcl_df.reindex(columns=headers)

    for i, row in dcl_df_to_write.iterrows():
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws_result.cell(row=start_row + i, column=col_idx)
            if not isinstance(cell, MergedCell):
                val = row.get(col_name, "")
                if isinstance(val, (pd.Series, list)):
                    val = val.iloc[0] if hasattr(val, "iloc") else val[0]
                if pd.isna(val):
                    val = ""
                cell.value = val

    # 写入PartsCoverage（dat_data["data"]，同理按表头名称写入，表头在第10行，数据从A11开始）
    if dat_data and "data" in dat_data:
        dat_df = dat_data["data"]
        parts_header_row = 10
        parts_start_row = 11
        parts_headers = [cell.value for cell in ws_parts[parts_header_row]]

        # 确保Description在表头中
        if "Description" not in parts_headers:
            parts_headers.append("Description")

                # 只保留有Description的行，并重排No.为连续自然数
        dat_df = dat_df[dat_df["Description"].notna() & (dat_df["Description"] != "")]
        dat_df = dat_df.reset_index(drop=True)
        if "No." in dat_df.columns:
            dat_df["No."] = range(1, len(dat_df) + 1)

        # 自动生成Remark列
        def gen_remark(row):
            if row.get("Testable") == "N":
                return "No test point"
            elif row.get("Testable") == "L":
                comps = str(row.get("Components", ""))
                if "/" in comps:
                    return f"it is in parallel with {comps.split('/')[-1].strip()}"
                elif "," in comps:
                    return f"it is in parallel with {comps.split(',')[-1].strip()}"
                else:
                    return "it is in parallel with"
            else:
                return row.get("Remark", "")
        dat_df["Remark"] = dat_df.apply(gen_remark, axis=1)

        # 保证dat_df列顺序与表头一致，只保留表头列
        dat_df_to_write = dat_df.reindex(columns=parts_headers)

        for i, row in dat_df_to_write.iterrows():
            for col_idx, col_name in enumerate(parts_headers, start=1):
                cell = ws_parts.cell(row=parts_start_row + i, column=col_idx)
                if not isinstance(cell, MergedCell):
                    val = row.get(col_name, "")
                    if isinstance(val, (pd.Series, list)):
                        val = val.iloc[0] if hasattr(val, "iloc") else val[0]
                    if pd.isna(val):
                        val = ""
                    cell.value = val

    # 写入dat_data的board_name和test_time到PartsCoverage的指定单元格
    if dat_data:
        ws_parts['H4'] = dat_data.get("board_name", "")
        ws_parts['H7'] = dat_data.get("test_time", "")
        
        coverage = dat_data.get("coverage", None)
        if coverage is not None:
            ws_parts['C7'] = "{:.2%}".format(coverage)
            
     # 给Test Result sheet加边框
    if not dcl_df_to_write.empty:
        min_row = start_row
        max_row = start_row + len(dcl_df_to_write) - 1
        min_col = 1
        max_col = len(headers)
        add_full_border(ws_result, min_row, max_row, min_col, max_col)

    # 给PartsCoverage sheet加边框
    if dat_data and "data" in dat_data and not dat_df_to_write.empty:
        min_row = parts_start_row
        max_row = parts_start_row + len(dat_df_to_write) - 1
        min_col = 1
        max_col = len(parts_headers)
        add_full_border(ws_parts, min_row, max_row, min_col, max_col)
    
    # 给Test Result sheet加边框
    if not dcl_df_to_write.empty:
        min_row = start_row
        max_row = start_row + len(dcl_df_to_write) - 1
        min_col = 1
        max_col = len(headers)
        add_full_border(ws_result, min_row, max_row, min_col, max_col)
        # 居中显示
        for row in ws_result.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # 给PartsCoverage sheet加边框
    if dat_data and "data" in dat_data and not dat_df_to_write.empty:
        min_row = parts_start_row
        max_row = parts_start_row + len(dat_df_to_write) - 1
        min_col = 1
        max_col = len(parts_headers)
        add_full_border(ws_parts, min_row, max_row, min_col, max_col)
        # 居中显示
        for row in ws_parts.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
   
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
